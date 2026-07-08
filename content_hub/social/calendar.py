"""social.calendar — the Social Calendar spreadsheet is the source of truth.

Reads the local working copy of Ghedee_Social_Calendar_<id>_v<n>.xlsx, turns each
in-scope Draft row into media-generation job(s), and (after generation + upload)
writes the Drive link and cost back into that row as a real hyperlink.

Columns are located by header name (row 1), not fixed index, so adding the new
'Slides' column — or any reordering — doesn't break the reader.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from . import rules
from ..core import config

# Carousel prompt parsing: one prompt per slide, each marked "Slide N:" (a dash — – -
# is also accepted). Optional per-slide 'On-image text: "…"' and a trailing global
# Style/Palette clause that applies to every slide.
_SLIDE_MARK = re.compile(r"Slide\s+\d+\s*[—–:\-]\s*", re.IGNORECASE)
_STYLE_TAIL = re.compile(r"\b(?:Style|Palette)\s*:", re.IGNORECASE)
_ONIMG = re.compile(r"On-image text\s*:", re.IGNORECASE)
_QUOTES = "“”\"'‘’"

# Carousel slides are painted clean and the wording is stamped on afterward, so the
# MODEL must render none. Image models tend to invent titles/subtitles, so this is
# deliberately forceful and repeated in both the prompt and the negatives.
_NO_TEXT_DIRECTIVE = (
    " CRITICAL: This image must contain absolutely NO text of any kind — no words, "
    "letters, captions, titles, subtitles, labels, numbers, signage, handwriting or "
    "typography anywhere in the frame. Render only the scene. Keep the lower third as "
    "clean, unobstructed negative space (text is added separately afterward).")
_NO_TEXT_NEGATIVE = (", any text, words, letters, captions, titles, subtitles, labels, "
                     "numbers, signage, handwriting, typography, watermark")


def _balance_quotes(text: str | None) -> str | None:
    """Append matching closing curly quotes for any unmatched opening ones, so an
    overlay like  The voice that calls you ‘not enough.  reads  …‘not enough.’"""
    if not text:
        return text
    for open_c, close_c in (("‘", "’"), ("“", "”")):
        diff = text.count(open_c) - text.count(close_c)
        if diff > 0:
            text += close_c * diff
    return text


def parse_carousel_prompt(prompt: str) -> list[dict]:
    """Split a multi-slide carousel prompt into per-slide {desc, text}.

    ``desc`` is the slide's visual direction with the global Style/Palette clause
    appended and the on-image-text directive removed (so the model paints clean
    artwork); ``text`` is the exact wording to overlay, or None. Returns [] if the
    prompt has no ``Slide N`` markers (caller falls back to the whole prompt)."""
    prompt = prompt or ""
    mtail = _STYLE_TAIL.search(prompt)
    body = prompt[:mtail.start()] if mtail else prompt
    tail = prompt[mtail.start():].strip() if mtail else ""

    marks = list(_SLIDE_MARK.finditer(body))
    slides = []
    for i, m in enumerate(marks):
        end = marks[i + 1].start() if i + 1 < len(marks) else len(body)
        seg = body[m.end():end].strip()
        om = _ONIMG.search(seg)
        if om:
            desc = seg[:om.start()].strip()
            text = _balance_quotes(seg[om.end():].strip().strip(_QUOTES).strip()) or None
        else:
            desc, text = seg, None
        slides.append({"desc": (f"{desc} {tail}".strip() if tail else desc), "text": text})
    return slides

# canonical field -> accepted header spellings (normalised: lower, single-spaced)
_HEADER_ALIASES = {
    "row_id": {"row id"},
    "date": {"date"},
    "day": {"day"},
    "time": {"time (pt)", "time (et)", "time"},
    "platform": {"platform"},
    "pillar": {"content pillar", "pillar"},
    "format": {"format"},
    "hook": {"headline", "hook / headline", "hook/headline", "hook"},
    "caption": {"caption", "caption (full copy)"},
    "hashtags": {"first-comment hashtags (ig)", "first-comment hashtags", "hashtags"},
    "visual_type": {"visual type"},
    "visual_direction": {"visual direction"},
    # The Prompt column is what actually drives image/video generation.
    "prompt": {"prompt"},
    "ai_model": {"ai model"},
    "est_cost": {"est. cost (usd)", "est cost (usd)", "est. cost", "est cost"},
    "asset_link": {"generated asset link", "generated asset link (drive)"},
    # A human-picked asset to use INSTEAD of generating: when set, generate copies it
    # into the target location rather than calling the image model. (Header: "Created
    # Asset Link"; internal field name kept as selected_asset to limit churn.)
    "selected_asset": {"created asset link", "selected asset link (drive)",
                       "selected asset link", "selected asset"},
    "status": {"status"},
    "notes": {"your notes", "notes"},
    "slides": {"slides", "slide count", "# slides"},
    "carousel_group": {"carousel group", "group"},
}

LINK_FONT_COLOR = "0563C1"  # Excel's default hyperlink blue


# --- new-calendar shell -----------------------------------------------------
# The full column set the team works in (matches the living calendar's layout).
# The reader locates columns by header name, so this ordering is the human-friendly
# default Cowork starts from, not a hard contract. Status leads (and is frozen with
# Row ID) so a reviewer always sees each post's id + status while scrolling.
SHELL_HEADERS = [
    "Status", "Row ID", "Date", "Day", "Time (PT)", "Content Pillar", "External Link",
    "Platform", "Format", "Slides", "Headline", "Caption",
    "First-comment Hashtags (IG)", "Visual Type", "Visual Direction", "Prompt",
    "AI Model", "Est. Cost (USD)", "Generated Asset Link", "Created Asset Link",
    "Your Notes", "Revision (Claude)",
]
SHELL_COL_WIDTHS = [13, 14, 6, 10, 26, 14, 16, 22, 28, 10, 30, 18, 50, 26, 26, 40, 16,
                    30, 13, 13, 13, 13]
FROZEN_COLUMNS = 2  # Status + Row ID stay put when scrolling horizontally

# Constrained columns -> their allowed values (rendered as dropdowns in the shell so a
# new calendar enforces them; Platform + Format together describe the entry, e.g.
# "Instagram Reel"). Visual Type is derived from Format (Carousel -> AI text-to-carousel).
PLATFORM_VALUES = ["Instagram", "Facebook", "Tiktok"]
FORMAT_VALUES = ["Post", "Reel", "Carousel"]
STATUS_VALUES = ["Draft", "Awaiting Asset", "Wiah Review", "Approved"]
HEADER_FILL = "FF1B3A2D"          # deep forest green (opaque ARGB)
HEADER_FONT_COLOR = "FFF7F2E8"    # ivory
HEADER_BORDER_COLOR = "FF7B9E87"  # sage
DEFAULT_TAB_TITLE = "Calendar"  # kept generic — the id may be a range or single day

# Status -> cell fill (light tints, readable with dark text). Kept in sync with the
# preview's status palette (content_hub/social/preview.py) and the Apps Script
# conditional-formatting installer (docs/appscript_status_endpoint.gs). Blank = no fill.
STATUS_FILLS = {
    "Draft": "FDECB0",           # yellow
    "Approved": "BCE8C8",        # green
    "Awaiting Asset": "DDE0E4",  # gray
    "Wiah Review": "E0D2F7",     # purple
}
STATUS_OTHER_FILL = "F7C7C2"     # any other non-blank status -> red


def add_status_conditional_formatting(ws, headers=SHELL_HEADERS, last_row: int = 1000):
    """Colour the Status column by value via native conditional formatting, so a status
    set anywhere — the preview OR a direct edit — is coloured automatically, matching the
    preview palette. Rules cover the four known statuses plus a catch-all (any other
    non-blank value -> red)."""
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
    col = get_column_letter(headers.index("Status") + 1)
    rng = f"{col}2:{col}{last_row}"
    for status, rgb in STATUS_FILLS.items():
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="equal", formula=[f'"{status}"'],
            fill=PatternFill("solid", fgColor=rgb)))
    known = ",".join(f'${col}2<>"{s}"' for s in STATUS_FILLS)
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'AND(${col}2<>"",{known})'],
        fill=PatternFill("solid", fgColor=STATUS_OTHER_FILL)))


def add_dropdowns(ws, headers=SHELL_HEADERS, last_row: int = 1000):
    """Data-validation dropdowns for the constrained columns (Platform, Format, Visual
    Type, Status), so a new calendar enforces the allowed values. Google Sheets imports
    these as native dropdowns on the living sheet."""
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    choices = {
        "Platform": PLATFORM_VALUES,
        "Format": FORMAT_VALUES,
        "Visual Type": [rules.VT_IMAGE, rules.VT_VIDEO, rules.VT_CAROUSEL, rules.VT_RECORDED],
        "Status": STATUS_VALUES,
    }
    for header, values in choices.items():
        if header not in headers:
            continue
        col = get_column_letter(headers.index(header) + 1)
        dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"',
                            allow_blank=True)  # arrow shows (showDropDown left unset)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{last_row}")


def build_shell_workbook(tab_title: str = DEFAULT_TAB_TITLE):
    """A header-only calendar workbook matching the team's column layout, styled to
    match the living calendar (deep-green header, ivory bold text, sage borders). The
    header row and the first two columns (Status, Row ID) are frozen, and the Status
    column is colour-coded by value. Cowork fills the Draft rows in; the reader locates
    columns by header name, so the exact ordering here is a convenience, not a contract."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tab_title or DEFAULT_TAB_TITLE
    fill = PatternFill("solid", fgColor=HEADER_FILL)
    font = Font(name="Calibri", size=11, bold=True, color=HEADER_FONT_COLOR)
    side = Side(style="thin", color=HEADER_BORDER_COLOR)
    border = Border(left=side, right=side, top=side, bottom=side)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, (head, width) in enumerate(zip(SHELL_HEADERS, SHELL_COL_WIDTHS), start=1):
        cell = ws.cell(1, i, head)
        cell.fill, cell.font, cell.border, cell.alignment = fill, font, border, align
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 46.25
    # Freeze the header row and the first FROZEN_COLUMNS columns (cell = first scrollable).
    ws.freeze_panes = f"{get_column_letter(FROZEN_COLUMNS + 1)}2"
    add_status_conditional_formatting(ws)
    add_dropdowns(ws)
    return wb


def new_shell_bytes(tab_title: str = DEFAULT_TAB_TITLE) -> bytes:
    """The shell workbook as .xlsx bytes (for uploading as a Google Sheet and/or
    writing a local working copy)."""
    import io
    buf = io.BytesIO()
    build_shell_workbook(tab_title).save(buf)
    return buf.getvalue()


def _norm(s) -> str:
    return " ".join(str(s or "").strip().lower().split())


@dataclass
class RowJob:
    """One calendar row's generation plan."""
    row_index: int                 # 1-based worksheet row
    row_id: str
    platform: str
    pillar: str
    hook: str
    status: str
    visual_type: str
    fmt: str
    prompt: str
    model: str
    plan: rules.VisualPlan
    assets: list[dict] = field(default_factory=list)  # core.media asset dicts
    group: str | None = None       # carousel group folder name
    existing_link: str | None = None
    selected_link: str | None = None  # human-picked asset to copy instead of generating
    skip_reason: str = ""          # non-empty => don't generate (status/visual/etc.)
    # display fields (for the review preview; not used by generation)
    date: str = ""
    day: str = ""
    caption: str = ""
    hashtags: str = ""
    notes: str = ""
    visual_direction: str = ""  # human art-direction notes; the Prompt column generates

    @property
    def in_scope(self) -> bool:
        return not self.skip_reason


class Calendar:
    """A loaded calendar workbook + its header map. Read jobs, then write links back."""

    def __init__(self, source):
        """``source`` is a filesystem path, or a file-like object / BytesIO (e.g. an
        xlsx downloaded from Drive for a read-only preview). ``path`` is None for the
        latter, so save() is a no-op target and must be given an explicit path."""
        import openpyxl
        if isinstance(source, (str, Path)):
            self.path = Path(source)
            self.wb = openpyxl.load_workbook(self.path)
        else:
            self.path = None
            self.wb = openpyxl.load_workbook(source)
        self.ws = self._pick_sheet()
        self.cols = self._map_headers()

    def _pick_sheet(self):
        for ws in self.wb.worksheets:
            if "calendar" in ws.title.lower():
                return ws
        return self.wb.worksheets[0]

    def _map_headers(self) -> dict[str, int]:
        header_to_field = {h: fld for fld, hs in _HEADER_ALIASES.items() for h in hs}
        cols: dict[str, int] = {}
        for c in range(1, self.ws.max_column + 1):
            fld = header_to_field.get(_norm(self.ws.cell(1, c).value))
            if fld and fld not in cols:
                cols[fld] = c
        missing = {"row_id", "visual_type", "prompt", "status", "asset_link"} - cols.keys()
        if missing:
            raise ValueError(f"calendar is missing required column(s): {sorted(missing)}")
        return cols

    def _get(self, row: int, field_name: str):
        c = self.cols.get(field_name)
        return self.ws.cell(row, c).value if c else None

    def _cell_link(self, row: int, field_name: str) -> str | None:
        """A URL from a link-bearing cell: the hyperlink target if the cell is a real
        hyperlink (display text may be a friendly label), else the plain cell value.
        Used for the human-entered Selected Asset Link, which may be pasted either way."""
        c = self.cols.get(field_name)
        if not c:
            return None
        cell = self.ws.cell(row, c)
        if cell.hyperlink and cell.hyperlink.target:
            return str(cell.hyperlink.target).strip()
        v = cell.value
        s = str(v).strip() if v is not None else ""
        return s or None

    @staticmethod
    def _fmt_date(v) -> str:
        """Date cell -> 'YYYY-MM-DD' (openpyxl may hand back a datetime or a string)."""
        if v is None:
            return ""
        if hasattr(v, "strftime"):
            return v.strftime("%Y-%m-%d")
        return str(v).strip()

    # --- read -----------------------------------------------------------------
    def read_jobs(self) -> list[RowJob]:
        jobs: list[RowJob] = []
        for r in range(2, self.ws.max_row + 1):
            row_id = self._get(r, "row_id")
            if not row_id:
                continue  # blank/spacer row
            job = self._build_job(r, str(row_id).strip())
            jobs.append(job)
        return jobs

    def _build_job(self, r: int, row_id: str) -> RowJob:
        status = str(self._get(r, "status") or "").strip()
        visual_type = str(self._get(r, "visual_type") or "").strip()
        fmt = str(self._get(r, "format") or "").strip()
        prompt = str(self._get(r, "prompt") or "").strip()
        platform = str(self._get(r, "platform") or "").strip()
        pillar = str(self._get(r, "pillar") or "").strip()
        hook = str(self._get(r, "hook") or "").strip()
        model_cell = str(self._get(r, "ai_model") or "").strip()
        plan = rules.plan_visual(visual_type, fmt)

        job = RowJob(row_index=r, row_id=row_id, platform=platform, pillar=pillar,
                     hook=hook, status=status, visual_type=visual_type, fmt=fmt,
                     prompt=prompt, model=model_cell, plan=plan,
                     existing_link=self._get(r, "asset_link"),
                     selected_link=self._cell_link(r, "selected_asset"),
                     date=self._fmt_date(self._get(r, "date")),
                     day=str(self._get(r, "day") or "").strip(),
                     caption=str(self._get(r, "caption") or "").strip(),
                     hashtags=str(self._get(r, "hashtags") or "").strip(),
                     notes=str(self._get(r, "notes") or "").strip(),
                     visual_direction=str(self._get(r, "visual_direction") or "").strip())

        # Only Draft rows are ever generated (per the workflow spec).
        if status.lower() != "draft":
            job.skip_reason = f"status is {status!r}, not Draft"
            return job
        # Recorded-Wiah clips are never AI-generated; they're sourced only from the
        # Selected Asset column (copied into 01_Wiah Videos). Without one there's
        # nothing to do yet — the film hasn't been uploaded.
        if plan.recorded:
            if not job.selected_link:
                job.skip_reason = plan.reason
                return job
            job.assets = self._assets_for(job)
            return job
        if not plan.generate:
            job.skip_reason = plan.reason
            return job
        # A carousel needs a Slides count. Its slides come from either the Prompt (one
        # "Slide N:" prompt per slide, count must match Slides) OR — when a Created Asset
        # Link is set — the images in that folder (count checked against Slides at generate
        # time). Handled here so a folder-sourced carousel isn't rejected for an empty Prompt.
        if job.plan.kind == "carousel":
            want = self._slide_count(job)  # 0 when the Slides cell is blank/invalid
            if not want:
                job.skip_reason = "carousel needs a Slides value (the number of slides)"
                return job
            if not job.selected_link:
                got = len(parse_carousel_prompt(prompt))
                if got == 0:
                    job.skip_reason = ("carousel Prompt has no per-slide prompts — write one "
                                       "per slide as 'Slide 1: …', 'Slide 2: …', …")
                    return job
                if want != got:
                    job.skip_reason = (f"carousel Prompt has {got} slide prompt(s) but the "
                                       f"Slides column says {want} — they must match")
                    return job
            job.assets = self._assets_for(job)
            return job
        if not prompt:
            job.skip_reason = "no prompt in Prompt column"
            return job

        job.assets = self._assets_for(job)
        return job

    def _assets_for(self, job: RowJob) -> list[dict]:
        neg = config.DEFAULT_NEGATIVE_PROMPT
        plat_short = rules.platform_short(job.platform, job.row_id)
        pillar_slug = rules.slugify(job.hook or job.pillar)
        model = job.model or (config.DEFAULT_VIDEO_MODEL if job.plan.kind == "video"
                              else config.DEFAULT_IMAGE_MODEL)

        if job.plan.kind == "carousel":
            # One asset per slide. Slides come from the Prompt's "Slide N:" prompts, or —
            # when a Created Asset Link is set — from the images in that folder (filled at
            # generate time), so build slide placeholders by the Slides count.
            group = str(self._get(job.row_index, "carousel_group") or "").strip() \
                or f"{job.row_id}_{pillar_slug}"
            job.group = group
            if job.selected_link:
                return [{"id": f"slide-{i}", "type": "image", "group": group,
                         "usage": f"{plat_short.lower()}-carousel"}
                        for i in range(1, self._slide_count(job) + 1)]
            assets = []
            for n, seg in enumerate(parse_carousel_prompt(job.prompt), start=1):
                desc = seg["desc"] + _NO_TEXT_DIRECTIVE
                assets.append({
                    "id": f"slide-{n}", "type": "image",
                    "prompt": desc,  # per-slide art direction, text-free
                    "negative_prompt": neg + _NO_TEXT_NEGATIVE,
                    "aspect_ratio": job.plan.aspect_ratio,
                    "model": model, "group": group,
                    "usage": f"{plat_short.lower()}-carousel",
                    "overlay_text": seg.get("text"),  # exact words stamped after render
                })
            return assets

        stem = f"{job.row_id}_{plat_short}_{pillar_slug}"
        asset = {"id": stem, "type": job.plan.kind, "prompt": job.prompt,
                 "negative_prompt": neg, "aspect_ratio": job.plan.aspect_ratio,
                 "model": model, "usage": f"{plat_short.lower()}-{job.plan.kind}"}
        if job.plan.kind == "video":
            # Veo 16:9 hero clips: default 720p/6s keeps cost/spec valid (1080p needs 8s).
            asset["resolution"] = "720p"
            asset["duration_seconds"] = 6
        return [asset]

    def _slide_count(self, job: RowJob) -> int:
        """The Slides column value, or 0 when blank/invalid (a carousel row with 0 is a
        validation error — Slides is required)."""
        raw = self._get(job.row_index, "slides")
        try:
            return int(raw)
        except (TypeError, ValueError):
            return 0

    # --- write ----------------------------------------------------------------
    @staticmethod
    def is_blank(v) -> bool:
        """True if a cell value counts as empty (None or a whitespace-only string)."""
        return v is None or (isinstance(v, str) and not v.strip())

    def write_result(self, row_index: int, link: str | None = None,
                     cost: float | str | None = None, model: str | None = None,
                     overwrite: bool = True) -> None:
        """Write the Generated Asset Link cell, and/or the cost + AI Model cells.

        A ``link`` that is a URL is written as a real, styled hyperlink; any other
        string (e.g. a ``Failed`` status) is written as plain text with no hyperlink
        or link styling, so a status never masquerades as a clickable link. ``model``
        records the model that actually produced the asset, so a per-run override is
        reflected in the sheet.

        ``overwrite`` is True for a live run (always writes). In a non-live rehearsal
        (mock/dry-run) it is False, so a cell is written only when it's currently blank
        — a rehearsal's estimate/placeholder never clobbers a real value."""
        from openpyxl.styles import Font

        def _col(field: str) -> int | None:
            c = self.cols.get(field)
            if not c or (not overwrite and not self.is_blank(self.ws.cell(row_index, c).value)):
                return None
            return c

        if link is not None:
            c = _col("asset_link")
            if c:
                cell = self.ws.cell(row_index, c)
                cell.value = link
                if isinstance(link, str) and link.lower().startswith(("http://", "https://")):
                    cell.hyperlink = link
                    cell.font = Font(color=LINK_FONT_COLOR, underline="single")
                else:
                    cell.hyperlink = None
                    cell.font = Font()
        if cost is not None:
            c = _col("est_cost")
            if c:
                self.ws.cell(row_index, c).value = cost
        if model is not None:
            c = _col("ai_model")
            if c:
                self.ws.cell(row_index, c).value = model

    NOTE_MARKER = "[auto]"

    @classmethod
    def merged_note(cls, existing, text: str) -> str:
        """Merge an automated note into an existing Notes cell: preserve human-written
        lines, replace the single ``[auto]`` line (so re-runs don't stack), and drop it
        entirely when ``text`` is empty. Shared by the openpyxl and Sheets writers."""
        kept = [ln for ln in str(existing or "").splitlines()
                if not ln.startswith(cls.NOTE_MARKER)]
        if text:
            kept.append(f"{cls.NOTE_MARKER} {text}")
        return "\n".join(kept).strip()

    def write_note(self, row_index: int, text: str) -> None:
        """Record an automated note in the Notes column, preserving human notes."""
        if "notes" not in self.cols:
            return
        cell = self.ws.cell(row_index, self.cols["notes"])
        cell.value = self.merged_note(cell.value, text) or None

    def save(self, path: Path | None = None) -> Path:
        dest = Path(path) if path else self.path
        self.wb.save(dest)
        return dest
