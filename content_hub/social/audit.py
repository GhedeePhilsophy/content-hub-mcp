"""social.audit — compliance audit of a calendar's posts vs. the canonical standards.

Answers, per post: *is this delivered asset and this caption fit for where it's going to
be published?* Each row is classified to a platform × post-type (see [specs]), then its
real Drive asset is downloaded and measured — **aspect ratio, resolution, file size,
video duration** — and its caption checked for **length / hashtag count / fold / placement /
links**. Row-level **readiness** (an Approved row with no asset, an empty caption, missing
schedule fields) and **link-sharing** are checked too, plus duplicate hashtags / assets.

Findings are returned as a structured report and (in ``live`` mode) written back into the
sheet's **Audit Status** (PASS / WARN / FAIL, colour-coded) and **Audit Note** (the reasons,
blank when PASS) columns. Read-only w.r.t. assets: nothing is generated, moved, deleted,
re-permissioned, or billed.

Modes (same three-mode contract as the rest of the workflow):
  dry-run  classify + caption/readiness checks only; assets are NOT downloaded, nothing written.
  mock     full read-only inspection (downloads + measures) but does NOT write to the sheet.
  live     full inspection AND writes Audit Status + Audit Note (ensuring/colouring the columns).

stdout is never touched (MCP channel) — progress goes through ``emit`` (stderr); the function
returns a structured dict. The standards live in [specs] so generation and the audit agree.
"""

from __future__ import annotations

import io
import re
import sys
from dataclasses import asdict, dataclass, field

from . import specs
from ..core import config
from ..core.drive import FOLDER_MIME, file_id_from_link

# verdict severity ordering (worst wins for a row's overall verdict; FAIL beats WARN)
_SEVERITY = {"NA": 0, "PASS": 1, "WARN": 2, "FAIL": 3}
AUDIT_STATUS_VALUES = ("PASS", "WARN", "FAIL")

KNOWN_RATIOS = {
    "1:1": 1.0, "4:5": 0.8, "5:4": 1.25, "9:16": 0.5625, "16:9": 16 / 9,
    "1.91:1": 1.91, "2:3": 2 / 3, "3:2": 1.5, "3:4": 0.75, "4:3": 4 / 3,
}
ASPECT_TOLERANCE = 0.04


def _stderr(msg: str, **_k) -> None:
    print(msg, file=sys.stderr)


# --- pure helpers (property-tested) ----------------------------------------
def parse_aspect(label: str) -> float | None:
    try:
        w, h = label.split(":")
        return float(w) / float(h)
    except (ValueError, ZeroDivisionError):
        return None


def measured_aspect_label(w: int, h: int) -> str:
    if not w or not h:
        return "?"
    val = w / h
    best, bestdiff = None, None
    for label, rv in KNOWN_RATIOS.items():
        diff = abs(val - rv) / rv
        if bestdiff is None or diff < bestdiff:
            best, bestdiff = label, diff
    if bestdiff is not None and bestdiff <= ASPECT_TOLERANCE:
        return best
    from math import gcd
    g = gcd(w, h) or 1
    return f"{w // g}:{h // g}"


def aspect_matches(w: int, h: int, accepted: tuple[str, ...],
                   tol: float = ASPECT_TOLERANCE) -> bool:
    if not w or not h:
        return False
    val = w / h
    for label in accepted:
        rv = parse_aspect(label)
        if rv and abs(val - rv) / rv <= tol:
            return True
    return False


_HASHTAG_RE = re.compile(r"(?<!\w)#\w+")
_URL_RE = re.compile(r"(https?://|www\.)\S+", re.IGNORECASE)


def count_hashtags(*texts: str | None) -> int:
    return sum(len(_HASHTAG_RE.findall(t or "")) for t in texts)


def find_urls(text: str | None) -> list[str]:
    return [m.group(0) for m in _URL_RE.finditer(text or "")]


def duplicate_asset_conflicts(a_platform: str, a_kind: str,
                              b_platform: str, b_kind: str) -> tuple[str, ...]:
    """Why two rows sharing one asset file is a finding — empty when it's fine.

    Reusing one asset across posts is legitimate and common (the same clip on a TikTok
    and on a Facebook Reel). It is only worth flagging when:

      ``same_platform``  both posts go to the same platform — the audience would see
                         the identical asset twice in one feed.
      ``kind_mismatch``  the rows disagree about what the asset *is* (an image post and
                         a video post cannot share one file; one of them is mis-typed).

    Kind comes from Visual Type via [rules.plan_visual], so 'AI text-to-video' and
    'Recorded video of Wiah' are the same kind — both are genuinely video.
    """
    reasons = []
    if specs.norm_platform(a_platform) == specs.norm_platform(b_platform):
        reasons.append("same_platform")
    if (a_kind or "") != (b_kind or ""):
        reasons.append("kind_mismatch")
    return tuple(reasons)


def duplicate_hashtags(*texts: str | None) -> list[str]:
    """Hashtags (case-insensitive) that appear more than once across the given fields."""
    seen: dict[str, int] = {}
    for t in texts:
        for tag in _HASHTAG_RE.findall(t or ""):
            seen[tag.lower()] = seen.get(tag.lower(), 0) + 1
    return [tag for tag, n in seen.items() if n > 1]


# --- check + row models ----------------------------------------------------
@dataclass
class Check:
    name: str
    verdict: str       # PASS | WARN | FAIL | NA
    detail: str


@dataclass
class RowAudit:
    row_id: str
    platform: str
    post_type: str
    fmt: str
    status: str
    visual_type: str = ""   # the row's Visual Type cell, verbatim
    kind: str = ""          # media kind derived from it: image | video | carousel | skip
    row_index: int = 0
    overall: str = "NA"
    checks: list[Check] = field(default_factory=list)
    issues: list[str] = field(default_factory=list)
    measured: dict = field(default_factory=dict)

    def add(self, name: str, verdict: str, detail: str) -> None:
        self.checks.append(Check(name, verdict, detail))

    def finalize(self) -> None:
        ranked = [c.verdict for c in self.checks] or ["NA"]
        self.overall = max(ranked, key=lambda v: _SEVERITY[v])

    def status_word(self) -> str:
        return self.overall if self.overall in AUDIT_STATUS_VALUES else ""

    def note_text(self) -> str:
        if self.overall in ("PASS", "NA"):
            return ""
        fails = [c.detail for c in self.checks if c.verdict == "FAIL"]
        warns = [c.detail for c in self.checks if c.verdict == "WARN"]
        parts = []
        if fails:
            parts.append("FAIL: " + "; ".join(fails))
        if warns:
            parts.append("WARN: " + "; ".join(warns))
        return " | ".join(parts)[:480]


# --- caption checks (no I/O) -----------------------------------------------
def audit_caption(row: RowAudit, caption: str, hashtags_field: str,
                  cspec: specs.CaptionSpec, platform: str) -> None:
    text = caption or ""
    n = len(text)
    if not text.strip():
        row.add("caption_length", "WARN", "caption is empty")
    elif n > cspec.caption_max:
        row.add("caption_length", "FAIL", f"caption {n} chars exceeds the {cspec.caption_max} cap")
    else:
        row.add("caption_length", "PASS",
                f"{n}/{cspec.caption_max} chars (first {cspec.fold} show before '…more')")
    row.measured["fold_preview"] = text[:cspec.fold]

    tags = count_hashtags(text, hashtags_field)
    row.measured["hashtags"] = tags
    if cspec.hashtag_max is not None and tags > cspec.hashtag_max:
        row.add("hashtags", "WARN", f"{tags} hashtags exceeds the {cspec.hashtag_max} max")
    elif tags:
        row.add("hashtags", "PASS", f"{tags} hashtags")

    p = specs.norm_platform(platform)
    # Links aren't clickable in Instagram/TikTok captions.
    if p in ("instagram", "tiktok") and find_urls(text):
        row.add("caption_links", "WARN",
                f"caption has a link — not clickable on {p.title()} (use link in bio)")
    # Instagram best practice: hashtags in the first comment, not the caption body.
    if p == "instagram" and count_hashtags(text) > 0:
        row.add("hashtag_placement", "WARN",
                "hashtags in the caption body — Instagram favours first-comment hashtags")
    dups = duplicate_hashtags(text, hashtags_field)
    if dups:
        row.add("hashtag_dupes", "WARN", "duplicate hashtags: " + ", ".join(dups[:6]))


# --- readiness checks (no I/O) ---------------------------------------------
def audit_readiness(row: RowAudit, job, time_val) -> None:
    """Coherence checks for a finished (Approved) row: it must actually be publishable.
    Drafts are works-in-progress, so these run only on Approved rows."""
    if (job.status or "").strip().lower() != "approved":
        return
    link = job.existing_link if isinstance(job.existing_link, str) else ""
    has_asset = (link.startswith("http")) or \
        (isinstance(job.selected_link, str) and job.selected_link.startswith("http"))
    if link.strip().lower() == "failed":
        row.add("readiness", "FAIL", "Approved but the asset generation is marked Failed")
    elif not has_asset and not job.plan.recorded:
        row.add("readiness", "FAIL", "Approved but has no asset link")
    if not (job.caption or "").strip():
        row.add("readiness_caption", "WARN", "Approved but the caption is empty")
    missing = [name for name, val in
               (("Platform", job.platform), ("Format", job.fmt),
                ("Date", job.date), ("Time", time_val))
               if not str(val or "").strip()]
    if missing:
        row.add("readiness_fields", "WARN", "missing " + ", ".join(missing))


# --- asset measurement (I/O) -----------------------------------------------
def _image_dims(raw: bytes) -> tuple[int, int] | None:
    try:
        from PIL import Image
        with Image.open(io.BytesIO(raw)) as im:
            return im.size
    except Exception:
        return None


def _video_meta(raw: bytes) -> tuple[int | None, int | None, float | None]:
    """(w, h, duration_seconds) of an mp4 via imageio+ffmpeg; Nones if unreadable."""
    import os
    import tempfile
    tmp = None
    try:
        import imageio
        tmp = tempfile.NamedTemporaryFile(suffix=".mp4", delete=False)
        tmp.write(raw)
        tmp.close()
        reader = imageio.get_reader(tmp.name)
        meta = reader.get_meta_data() or {}
        size = meta.get("size")
        if not size:
            frame = reader.get_data(0)
            size = (frame.shape[1], frame.shape[0])
        dur = meta.get("duration")
        if dur is None and meta.get("fps") and meta.get("nframes"):
            try:
                dur = float(meta["nframes"]) / float(meta["fps"])
            except (TypeError, ValueError, ZeroDivisionError):
                dur = None
        reader.close()
        return (int(size[0]), int(size[1]), float(dur) if dur else None)
    except Exception:
        return (None, None, None)
    finally:
        if tmp is not None:
            try:
                os.unlink(tmp.name)
            except OSError:
                pass


def _size_mb(meta: dict) -> float | None:
    try:
        return round(int(meta["size"]) / (1024 * 1024), 2)
    except (KeyError, TypeError, ValueError):
        return None


def _check_media(row: RowAudit, w: int | None, h: int | None, size_mb: float | None,
                 spec: specs.MediaSpec, *, duration: float | None = None,
                 prefix: str = "") -> None:
    if w and h:
        label = measured_aspect_label(w, h)
        row.measured[f"{prefix}dims"] = f"{w}x{h}"
        if aspect_matches(w, h, spec.accepted_aspects):
            row.add(f"{prefix}aspect", "PASS", f"{label} ({w}x{h})")
        else:
            row.add(f"{prefix}aspect", "FAIL",
                    f"{label} ({w}x{h}) not in accepted {list(spec.accepted_aspects)}")
        short_edge = min(w, h)
        if short_edge < spec.min_short_edge_px:
            row.add(f"{prefix}resolution", "WARN",
                    f"short edge {short_edge}px below the {spec.min_short_edge_px}px minimum")
        else:
            row.add(f"{prefix}resolution", "PASS", f"short edge {short_edge}px")
    else:
        row.add(f"{prefix}aspect", "NA", "could not read image/video dimensions")
    if size_mb is not None:
        row.measured[f"{prefix}size_mb"] = size_mb
        if size_mb > spec.max_file_mb:
            row.add(f"{prefix}file_size", "FAIL",
                    f"{size_mb} MB exceeds the {spec.max_file_mb} MB cap")
        else:
            row.add(f"{prefix}file_size", "PASS", f"{size_mb} MB")
    if duration is not None and (spec.min_seconds or spec.max_seconds):
        row.measured[f"{prefix}duration_s"] = round(duration, 1)
        if spec.max_seconds and duration > spec.max_seconds:
            row.add(f"{prefix}duration", "FAIL",
                    f"{duration:.0f}s exceeds the {spec.max_seconds:.0f}s max for this type")
        elif spec.min_seconds and duration < spec.min_seconds:
            row.add(f"{prefix}duration", "WARN",
                    f"{duration:.1f}s below the {spec.min_seconds:.0f}s minimum")
        else:
            row.add(f"{prefix}duration", "PASS", f"{duration:.0f}s")


def _audit_asset(drive, row: RowAudit, job, spec: specs.MediaSpec, expected_slides: int,
                 md5_sink: list, emit) -> None:
    """Download + measure the row's real asset(s). Any resolution failure is reported,
    never raised (one bad asset never kills the run)."""
    link = job.existing_link if isinstance(job.existing_link, str) else None
    if link and link.strip().lower() == "failed":
        row.add("asset", "NA", "generation marked Failed — nothing to audit")
        return
    if not (link and link.startswith("http")):
        link = job.selected_link
    if not (isinstance(link, str) and link.startswith("http")):
        if job.plan.recorded:
            row.add("asset", "NA", "recorded clip — awaiting footage")
        else:
            row.add("asset", "NA", "no asset link on the row")
        return
    fid = file_id_from_link(link)
    if not fid:
        row.add("asset", "NA", "asset link is not a parseable Drive URL")
        return

    try:
        if not drive.is_link_shared(fid):
            row.add("sharing", "WARN",
                    "asset is not shared 'anyone with link' — a scheduler fetch will fail")
        if job.plan.kind == "carousel":
            _audit_carousel(drive, row, fid, spec, expected_slides, emit)
            return
        meta = drive.get_file(fid, fields="id,name,mimeType,size,md5Checksum")
        if meta.get("mimeType") == FOLDER_MIME:
            _audit_carousel(drive, row, fid, spec, expected_slides, emit)
            return
        if meta.get("md5Checksum"):
            md5_sink.append((row, meta["md5Checksum"]))
        raw = drive.download_bytes(fid)
        is_video = job.plan.kind == "video" or meta.get("mimeType", "").startswith("video/")
        if is_video:
            w, h, dur = _video_meta(raw)
        else:
            dims = _image_dims(raw)
            w, h = dims if dims else (None, None)
            dur = None
        _check_media(row, w, h, _size_mb(meta), spec, duration=dur)
    except Exception as e:
        row.add("asset", "NA", f"could not fetch/measure asset: {str(e).splitlines()[0][:120]}")


def _audit_carousel(drive, row: RowAudit, folder_id: str, spec: specs.MediaSpec,
                    expected_slides: int, emit) -> None:
    slides = sorted(
        (f for f in drive.list_children(folder_id)
         if str(f.get("mimeType", "")).startswith("image/")),
        key=lambda f: f["name"])
    n = len(slides)
    row.measured["slides"] = n
    lo, hi = spec.carousel_min_slides, spec.carousel_max_slides
    if n == 0:
        row.add("slides", "NA", "carousel folder holds no images")
        return
    if (lo and n < lo) or (hi and n > hi):
        row.add("slides", "WARN", f"{n} slides outside the {lo}–{hi} range for this platform")
    else:
        row.add("slides", "PASS", f"{n} slides")
    # Slides column vs the actual folder count.
    if expected_slides and expected_slides != n:
        row.add("slides_count", "WARN",
                f"Slides column says {expected_slides} but the folder holds {n} images")
    # measure every slide; fold into one worst-case verdict per check + check ratio consistency
    worst = {"aspect": "PASS", "resolution": "PASS", "file_size": "PASS"}
    detail = {}
    labels = set()
    for f in slides:
        raw = drive.download_bytes(f["id"])
        dims = _image_dims(raw)
        w, h = dims if dims else (None, None)
        if w and h:
            labels.add(measured_aspect_label(w, h))
        sub = RowAudit(row.row_id, row.platform, row.post_type, row.fmt, row.status)
        _check_media(sub, w, h, _size_mb(f), spec)
        for c in sub.checks:
            if _SEVERITY[c.verdict] > _SEVERITY[worst.get(c.name, "NA")]:
                worst[c.name] = c.verdict
                detail[c.name] = f"slide {f['name']}: {c.detail}"
    for base, verdict in worst.items():
        row.add(base, verdict, detail.get(base, f"all {n} slides OK"))
    if len(labels) > 1:
        row.add("carousel_consistency", "WARN",
                f"slides mix aspect ratios {sorted(labels)} — Instagram crops to slide 1")


# --- living-sheet column setup (live only) ---------------------------------
_AUDIT_STATUS_HDR = "Audit Status"
_AUDIT_NOTE_HDR = "Audit Note"
# PASS green / WARN yellow / FAIL red (0-1 floats; same palette as the human Status column)
_CF_COLORS = {
    "PASS": {"red": 0.737, "green": 0.910, "blue": 0.784},
    "WARN": {"red": 0.992, "green": 0.925, "blue": 0.690},
    "FAIL": {"red": 0.969, "green": 0.780, "blue": 0.761},
}


def _ensure_audit_columns(sheets, sid: str, tab: str, cal, emit) -> tuple[int, int]:
    """Make sure the sheet has 'Audit Status' + 'Audit Note' columns (migrating a legacy
    single 'Audit Results' column into Audit Status), and install the PASS/WARN/FAIL dropdown
    + colour rules on Audit Status once. Returns their 1-based column indices."""
    from openpyxl.utils import get_column_letter
    hdr = {}
    for c in range(1, cal.ws.max_column + 1):
        v = cal.ws.cell(1, c).value
        if v:
            hdr[" ".join(str(v).strip().lower().split())] = c

    status_col = hdr.get("audit status")
    note_col = hdr.get("audit note")
    end = max([*hdr.values(), cal.ws.max_column]) if hdr else cal.ws.max_column
    header_writes: list[tuple[str, object]] = []
    newly_made = False

    if not status_col:
        status_col = hdr.get("audit results")  # migrate legacy single column
        if not status_col:
            end += 1
            status_col = end
        header_writes.append((f"'{tab}'!{get_column_letter(status_col)}1", _AUDIT_STATUS_HDR))
        newly_made = True
    if not note_col:
        end += 1
        note_col = end
        header_writes.append((f"'{tab}'!{get_column_letter(note_col)}1", _AUDIT_NOTE_HDR))

    if header_writes:
        sheets.batch_update(sid, header_writes)  # writing headers also expands the grid

    meta = sheets.sheet_meta(sid, tab.lower() if "calendar" in tab.lower() else tab)
    sheet_id = meta["sheetId"]
    col0 = status_col - 1
    rng = {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": 1000,
           "startColumnIndex": col0, "endColumnIndex": col0 + 1}
    requests: list[dict] = [{
        "setDataValidation": {
            "range": rng,
            "rule": {
                "condition": {"type": "ONE_OF_LIST",
                              "values": [{"userEnteredValue": v} for v in AUDIT_STATUS_VALUES]},
                "showCustomUi": True, "strict": False,
            },
        }
    }]
    # add colour rules only if this column has none yet (keeps re-runs idempotent)
    if col0 not in meta["cf_columns"]:
        for i, (val, color) in enumerate(_CF_COLORS.items()):
            requests.append({"addConditionalFormatRule": {"index": i, "rule": {
                "ranges": [rng],
                "booleanRule": {
                    "condition": {"type": "TEXT_EQ",
                                  "values": [{"userEnteredValue": val}]},
                    "format": {"backgroundColor": color},
                },
            }}})
    sheets.apply_requests(sid, requests)
    if newly_made:
        emit(f"audit: set up '{_AUDIT_STATUS_HDR}' (dropdown + colours) and '{_AUDIT_NOTE_HDR}'")
    return status_col, note_col


# --- orchestration ---------------------------------------------------------
def audit_calendar(calendar_id: str, *, mode: str = "dry-run",
                   statuses: tuple[str, ...] | None = None, emit=None) -> dict:
    """Audit a calendar's posts against the canonical per-platform specs.

    ``statuses`` optionally restricts which rows are audited (default None audits every row).
    Returns a structured report; in ``live`` mode also writes each row's Audit Status
    (PASS/WARN/FAIL, colour-coded) and Audit Note (reasons, blank when PASS).
    """
    emit = emit or _stderr
    if mode not in ("dry-run", "mock", "live"):
        raise ValueError(f"mode must be dry-run|mock|live, got {mode!r}")
    download = mode in ("mock", "live")
    write = mode == "live"

    from . import edit_ops
    drive, sid, tab, cal, sheet_link = edit_ops._load_live(calendar_id)

    rows: list[RowAudit] = []
    md5_sink: list = []
    jobs = [j for j in cal.read_jobs() if j.row_id]
    emit(f"audit: {calendar_id} [{mode}] — {len(jobs)} rows"
         + (f", statuses={list(statuses)}" if statuses else ""))

    for job in jobs:
        if statuses and job.status not in statuses:
            continue
        kind = job.plan.kind
        is_reel = (job.fmt or "").strip().lower() == "reel"
        post_type, issues = specs.classify(job.platform, job.fmt, kind, is_reel) \
            if kind in ("image", "video", "carousel") else (kind, [])
        row = RowAudit(job.row_id, job.platform, post_type, job.fmt, job.status,
                       visual_type=job.visual_type, kind=kind,
                       row_index=job.row_index, issues=issues)

        audit_caption(row, job.caption, job.hashtags, specs.caption_spec(job.platform),
                      job.platform)
        audit_readiness(row, job, cal._get(job.row_index, "time"))

        if kind not in ("image", "video", "carousel"):
            row.add("asset", "NA", job.plan.reason or f"visual type not audited ({kind})")
        elif not download:
            row.add("asset", "NA", "not inspected in dry-run (use --mode mock or live)")
        else:
            spec = specs.media_spec(job.platform, post_type)
            _audit_asset(drive, row, job, spec, cal._slide_count(job), md5_sink, emit)

        row.finalize()
        rows.append(row)

    # duplicate-asset post-pass (same Drive md5 reused across rows). Sharing an asset is
    # allowed — only the collisions in [duplicate_asset_conflicts] are reported.
    by_md5: dict[str, list[RowAudit]] = {}
    for r, md5 in md5_sink:
        by_md5.setdefault(md5, []).append(r)
    for md5, sharers in by_md5.items():
        if len(sharers) < 2:
            continue
        for r in sharers:
            clashes: dict[str, list[RowAudit]] = {}
            for other in sharers:
                if other is r:
                    continue
                for reason in duplicate_asset_conflicts(r.platform, r.kind,
                                                        other.platform, other.kind):
                    clashes.setdefault(reason, []).append(other)
            if not clashes:
                continue
            if clashes.get("same_platform"):
                ids = ", ".join(o.row_id for o in clashes["same_platform"])
                r.add("duplicate_asset", "WARN",
                      f"same asset file as {ids} — all on {r.platform or 'this platform'}")
            if clashes.get("kind_mismatch"):
                mine = r.visual_type or r.kind or "?"
                theirs = ", ".join(f"{o.row_id} ({o.visual_type or o.kind or '?'})"
                                   for o in clashes["kind_mismatch"])
                r.add("duplicate_asset_kind", "WARN",
                      f"same asset file as {theirs}, but this row is {mine} — "
                      "one asset cannot serve both")
            r.finalize()

    updated = 0
    if write:
        status_col, note_col = _ensure_audit_columns(sheets_client(), sid, tab, cal, emit)
        from openpyxl.utils import get_column_letter
        writebacks: list[tuple[str, object]] = []
        for r in rows:
            writebacks.append((f"'{tab}'!{get_column_letter(status_col)}{r.row_index}",
                               r.status_word()))
            writebacks.append((f"'{tab}'!{get_column_letter(note_col)}{r.row_index}",
                               r.note_text()))
        res = sheets_client().batch_update(sid, writebacks)
        updated = res.get("totalUpdatedCells", len(writebacks))
        emit(f"audit: wrote {updated} cell(s) to Audit Status / Audit Note")

    tally = {"pass": 0, "warn": 0, "fail": 0, "na": 0}
    for r in rows:
        tally[{"PASS": "pass", "WARN": "warn", "FAIL": "fail", "NA": "na"}[r.overall]] += 1
    emit(f"audit: {tally['fail']} fail / {tally['warn']} warn / {tally['pass']} pass "
         f"/ {tally['na']} n/a")

    return {
        "calendar_id": calendar_id,
        "mode": mode,
        "audited": len(rows),
        "summary": tally,
        "sheet_link": sheet_link,
        "wrote_back": updated if write else 0,
        "specs_verified": specs.VERIFIED,
        "rows": [_row_dict(r) for r in rows],
        "note": {
            "dry-run": "dry-run: assets not downloaded, sheet not written.",
            "mock": "mock: assets inspected read-only; sheet not written.",
            "live": "live: assets inspected; Audit Status + Audit Note written.",
        }[mode],
    }


_SHEETS = None


def sheets_client():
    """Lazily build one SheetsClient for the write-back + formatting calls."""
    global _SHEETS
    if _SHEETS is None:
        from ..core.sheets import SheetsClient
        _SHEETS = SheetsClient(config.credentials_path(), config.token_path())
    return _SHEETS


def _row_dict(r: RowAudit) -> dict:
    d = asdict(r)
    d["status"] = r.status_word()
    d["note"] = r.note_text()
    d["checks"] = [{"name": c["name"], "verdict": c["verdict"], "detail": c["detail"]}
                   for c in d["checks"]]
    return d
