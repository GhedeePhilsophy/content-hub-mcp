"""social.edit_ops — direct, schema-aware edits to the living Google Sheet.

Lets Cowork change the calendar in place — edit human-owned cells and append new
rows — without the download/modify/upload round-trip. Both operations reuse the same
Sheets-API writer that ``generate`` uses ([core.sheets.SheetsClient]), so an edit only
touches the exact cells named and never clobbers a teammate's concurrent change.

Three operations:

  edit_rows  change cells of existing rows, addressed by Row ID + column name.
  add_rows   append new rows in bulk (used to seed a fresh calendar).
  describe   report the sheet's columns + rows without changing it — the look-before-you-
             write companion, so a caller names real columns instead of guessing.

Guardrails (the reason this is a tool and not a raw Sheets connector):
  * Status is NEVER editable — approval is a human-only decision. add_rows may set the
    working states (Draft / Awaiting Asset) but never an approval state.
  * The machine-owned columns (Generated Asset Link / Est. Cost / AI Model) are written
    by generate; edit_rows refuses them unless force=True, add_rows refuses them outright.
  * Columns are resolved by header name (or a known alias); rows by the stable Row ID —
    so an edit survives reordering and can never land on the wrong row.
  * Constrained columns (Platform / Format / Visual Type) are validated against their
    allowed values.
"""

from __future__ import annotations

import io
import sys

from . import rules, sheet_ops
from .calendar import (
    Calendar, _HEADER_ALIASES, _norm,
    PLATFORM_VALUES, FORMAT_VALUES, STATUS_VALUES,
)
from ..core import config
from ..core.drive import DriveClient


def _stderr(msg: str, **_k) -> None:
    print(msg, file=sys.stderr)


# Canonical fields whose values generate owns — a human edit here would fight the
# generator, so they are force-gated (edit) / rejected (add).
MACHINE_FIELDS = {"asset_link", "est_cost", "ai_model"}
# Status is decided by a human reviewer, never by Cowork.
STATUS_FIELD = "status"
# The two review/approval states Cowork must never set (even when creating rows).
_APPROVAL_STATUSES = {"wiah review", "approved"}
# Constrained columns validated on write (field -> allowed values).
_CONSTRAINED = {
    "platform": PLATFORM_VALUES,
    "format": FORMAT_VALUES,
    "visual_type": [rules.VT_IMAGE, rules.VT_VIDEO, rules.VT_CAROUSEL, rules.VT_RECORDED],
}
# normalized header spelling -> canonical field (built from the reader's alias table).
_HEADER_TO_FIELD = {h: fld for fld, hs in _HEADER_ALIASES.items() for h in hs}


def _drive() -> DriveClient:
    return DriveClient(config.credentials_path(), config.token_path(), allow_interactive=False)


def _load_live(calendar_id: str):
    """Resolve the living Google Sheet and return (drive, sid, tab, Calendar, link).

    The Calendar is loaded from an .xlsx export purely to read the header map and the
    Row ID column — writes go back to the live sheet by A1 range via the Sheets API."""
    drive = _drive()
    docs = sheet_ops._docs_folder(drive, calendar_id)
    live = sheet_ops.find_live_sheet(drive, docs, calendar_id)
    if not live:
        raise FileNotFoundError(
            f"no live sheet '{sheet_ops.live_sheet_name(calendar_id)}' on Drive; "
            "run `social create` first.")
    sid = live["id"]
    cal = Calendar(io.BytesIO(drive.export_as_xlsx(sid)))
    return drive, sid, cal.ws.title, cal, live.get("webViewLink")


def _header_columns(cal) -> dict[str, int]:
    """normalized actual header text (row 1) -> 1-based column index."""
    cols: dict[str, int] = {}
    for c in range(1, cal.ws.max_column + 1):
        h = _norm(cal.ws.cell(1, c).value)
        if h and h not in cols:
            cols[h] = c
    return cols


def _resolve_column(cal, header_cols: dict[str, int], name: str):
    """(column_index, canonical_field) for a requested column name, or (None, None).

    Accepts either the exact header text or any known alias spelling."""
    field_by_col = {c: f for f, c in cal.cols.items()}
    norm = _norm(name)
    if norm in header_cols:                       # exact header match
        c = header_cols[norm]
        return c, field_by_col.get(c)
    fld = _HEADER_TO_FIELD.get(norm)              # alias -> field -> column
    if fld and fld in cal.cols:
        return cal.cols[fld], fld
    return None, None


def _a1(tab: str, col: int, row: int) -> str:
    from openpyxl.utils import get_column_letter
    return f"'{tab}'!{get_column_letter(col)}{row}"


def _validate_constrained(field: str | None, value) -> str | None:
    """Error string if ``value`` is not an allowed value for a constrained field, else None.
    A blank value is always allowed (clears the cell / leaves the dropdown empty)."""
    if field not in _CONSTRAINED:
        return None
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    allowed = _CONSTRAINED[field]
    if _norm(value) not in {_norm(a) for a in allowed}:
        return f"{field!r} must be one of {allowed}, got {value!r}"
    return None


def _rowid_index(cal) -> dict[str, int]:
    """normalized (lowercased, stripped) Row ID -> worksheet row index."""
    rid_col = cal.cols["row_id"]
    out: dict[str, int] = {}
    for r in range(2, cal.ws.max_row + 1):
        v = cal.ws.cell(r, rid_col).value
        key = str(v).strip().lower() if v is not None else ""
        if key:
            out.setdefault(key, r)
    return out


def _last_used_row(cal) -> int:
    """The last worksheet row that has a Row ID (new rows append after this)."""
    rid_col = cal.cols["row_id"]
    last = 1  # header
    for r in range(2, cal.ws.max_row + 1):
        v = cal.ws.cell(r, rid_col).value
        if v is not None and str(v).strip():
            last = r
    return last


# --- describe the sheet's shape ---------------------------------------------
def describe(calendar_id: str, *, emit=None) -> dict:
    """Report the living sheet's SHAPE — its columns and its rows — without changing it.

    This is the look-before-you-write companion to edit_rows / add_rows: it answers "what
    column names does this calendar actually have, and which Row IDs exist?" so a caller
    composes edits against the real header row instead of guessing (a guessed column name
    is rejected, and because the batch is all-or-nothing it takes every other edit with it).

    Returns, per column: the exact header text to pass as ``column``, its canonical field,
    and whether it is writable — ``editable`` false means edit_rows will refuse it, with
    ``why`` naming the guardrail ('status' is human-only; the machine-owned columns need
    force=True). ``allowed_values`` is set for the constrained columns.

    Returns, per row: the Row ID (pass it verbatim as ``row_id``), its worksheet row, and
    the human-owned summary fields (date / platform / format / status / headline), plus
    whether an asset link is present.

    Read-only: there is no ``mode`` because nothing is ever written and nothing is spent."""
    emit = emit or _stderr
    _drive_, _sid, tab, cal, link = _load_live(calendar_id)
    header_cols = _header_columns(cal)
    field_by_col = {c: f for f, c in cal.cols.items()}

    columns: list[dict] = []
    for c in range(1, cal.ws.max_column + 1):
        header = cal.ws.cell(1, c).value
        if header is None or not str(header).strip():
            continue
        field = field_by_col.get(c)
        info: dict = {"header": str(header).strip(), "field": field, "column": c,
                      "editable": True}
        if field == STATUS_FIELD:
            info["editable"] = False
            info["why"] = "approval is a human-only decision — set Status in the sheet"
        elif field in MACHINE_FIELDS:
            info["editable"] = False
            info["why"] = "written by generate — pass force=true to overwrite deliberately"
        elif field in ("audit_status", "audit_note"):
            info["why"] = "written by the audit each run — an edit here is overwritten"
        if field in _CONSTRAINED:
            info["allowed_values"] = list(_CONSTRAINED[field])
        columns.append(info)

    def _cell(r: int, field: str):
        c = cal.cols.get(field)
        v = cal.ws.cell(r, c).value if c else None
        return str(v).strip() if v is not None and str(v).strip() else None

    rows: list[dict] = []
    status_counts: dict[str, int] = {}
    for row_id, r in sorted(_rowid_index(cal).items(), key=lambda kv: kv[1]):
        status = _cell(r, "status")
        rows.append({
            "row_id": _cell(r, "row_id") or row_id,   # the sheet's own spelling
            "row": r,
            "date": _cell(r, "date"),
            "platform": _cell(r, "platform"),
            "format": _cell(r, "format"),
            "visual_type": _cell(r, "visual_type"),
            "status": status,
            "headline": _cell(r, "hook"),
            "has_asset": bool(_cell(r, "asset_link") or _cell(r, "selected_asset")),
        })
        status_counts[status or "(blank)"] = status_counts.get(status or "(blank)", 0) + 1

    emit(f"describe: {len(columns)} column(s), {len(rows)} row(s) in "
         f"{sheet_ops.live_sheet_name(calendar_id)}")
    return {"calendar_id": calendar_id, "tab": tab, "sheet_link": link,
            "columns": columns, "rows": rows,
            "row_count": len(rows), "status_counts": status_counts,
            "editable_columns": [c["header"] for c in columns if c["editable"]]}


# --- edit existing rows -----------------------------------------------------
def edit_rows(calendar_id: str, edits: list[dict], *, mode: str = "live",
              force: bool = False, emit=None) -> dict:
    """Change cells of existing calendar rows in the living Google Sheet, in place.

    ``edits`` is a list of ``{"row_id": ..., "column": ..., "value": ...}``. Columns are
    resolved by header name or alias; rows by Row ID. The whole batch is validated before
    anything is written: if ANY edit is invalid the run makes no changes and returns the
    errors (so a partial write can never leave the sheet half-edited).

    Status is rejected (approval is human-only). The machine-owned columns (Generated
    Asset Link / Est. Cost / AI Model) are rejected unless ``force`` is True.

    mode: 'dry-run' previews the resolved cell writes without touching the sheet; 'live'
    writes them. (There is no 'mock' here — nothing is generated and nothing is spent, so
    dry-run already is the safe rehearsal.)"""
    emit = emit or _stderr
    if mode not in ("dry-run", "live"):
        raise ValueError(f"mode must be 'dry-run' or 'live', got {mode!r}")
    if not edits:
        return {"calendar_id": calendar_id, "mode": mode, "planned": 0, "changes": [],
                "errors": [], "note": "no edits supplied"}

    drive, sid, tab, cal, link = _load_live(calendar_id)
    header_cols = _header_columns(cal)
    rows = _rowid_index(cal)

    updates: list[tuple[str, object]] = []
    changes: list[dict] = []
    errors: list[str] = []

    for i, e in enumerate(edits):
        row_id = str(e.get("row_id", "")).strip()
        column = str(e.get("column", "")).strip()
        value = e.get("value")
        where = f"edit[{i}] (row_id={row_id!r}, column={column!r})"
        if not row_id or not column:
            errors.append(f"{where}: both 'row_id' and 'column' are required")
            continue
        col, field = _resolve_column(cal, header_cols, column)
        if not col:
            errors.append(f"{where}: no column named {column!r} in the calendar")
            continue
        if field == STATUS_FIELD:
            errors.append(f"{where}: Status is not editable here — approval is a "
                          "human-only decision, set it in the sheet directly")
            continue
        if field in MACHINE_FIELDS and not force:
            errors.append(f"{where}: {column!r} is written by generate; pass force=true "
                          "to overwrite it deliberately")
            continue
        r = rows.get(row_id.lower())
        if not r:
            errors.append(f"{where}: no row with Row ID {row_id!r} in the calendar")
            continue
        cerr = _validate_constrained(field, value)
        if cerr:
            errors.append(f"{where}: {cerr}")
            continue
        updates.append((_a1(tab, col, r), value))
        changes.append({"row_id": row_id, "row": r, "column": column, "value": value})

    result = {"calendar_id": calendar_id, "mode": mode, "planned": len(changes),
              "changes": changes, "errors": errors, "sheet_link": link}
    if errors:
        result["note"] = (f"{len(errors)} invalid edit(s) — nothing was written. "
                          "Fix the errors and retry.")
        return result
    if mode == "dry-run":
        result["note"] = "dry-run: no cells were written."
        return result

    from ..core.sheets import SheetsClient
    sheets = SheetsClient(config.credentials_path(), config.token_path())
    res = sheets.batch_update(sid, updates)
    result["updated_cells"] = res.get("totalUpdatedCells", len(updates))
    emit(f"edit: wrote {result['updated_cells']} cell(s) to {sheet_ops.live_sheet_name(calendar_id)}")
    return result


# --- append new rows --------------------------------------------------------
def add_rows(calendar_id: str, rows: list[dict], *, mode: str = "live", emit=None) -> dict:
    """Append new rows to the living Google Sheet in bulk (seed a fresh calendar).

    ``rows`` is a list of column->value dicts, each keyed by header name or alias. Every
    row MUST carry a Row ID (the stable key); a Row ID already in the sheet, or repeated
    within the batch, is rejected. Rows are validated as a batch — if any row is invalid,
    nothing is written.

    Status defaults to 'Draft' (so generate will pick the row up); an explicit Status may
    only be a working state (Draft / Awaiting Asset) — an approval state is rejected. The
    machine-owned columns (Generated Asset Link / Est. Cost / AI Model) are rejected; they
    are filled by generate.

    mode: 'dry-run' previews the rows without writing; 'live' appends them."""
    emit = emit or _stderr
    if mode not in ("dry-run", "live"):
        raise ValueError(f"mode must be 'dry-run' or 'live', got {mode!r}")
    if not rows:
        return {"calendar_id": calendar_id, "mode": mode, "planned": 0, "rows": [],
                "errors": [], "note": "no rows supplied"}

    drive, sid, tab, cal, link = _load_live(calendar_id)
    header_cols = _header_columns(cal)
    existing_ids = set(_rowid_index(cal))
    start_row = _last_used_row(cal) + 1
    status_col = cal.cols.get("status")

    updates: list[tuple[str, object]] = []
    planned: list[dict] = []
    errors: list[str] = []
    seen: set[str] = set()

    for i, row in enumerate(rows):
        dest = start_row + len(planned)
        # locate this row's Row ID (accept header or alias spelling)
        row_id = ""
        for k, v in row.items():
            _, fld = _resolve_column(cal, header_cols, str(k))
            if fld == "row_id":
                row_id = str(v).strip()
                break
        where = f"row[{i}]"
        if not row_id:
            errors.append(f"{where}: a Row ID is required for every new row")
            continue
        key = row_id.lower()
        if key in existing_ids:
            errors.append(f"{where}: Row ID {row_id!r} already exists in the calendar")
            continue
        if key in seen:
            errors.append(f"{where}: Row ID {row_id!r} is repeated in this batch")
            continue
        seen.add(key)

        row_updates: list[tuple[str, object]] = []
        row_fields: set[str] = set()
        row_error = None
        for k, value in row.items():
            col, field = _resolve_column(cal, header_cols, str(k))
            if not col:
                row_error = f"{where}: no column named {str(k)!r} in the calendar"
                break
            if field in MACHINE_FIELDS:
                row_error = (f"{where}: {str(k)!r} is filled by generate and can't be set "
                             "when creating a row")
                break
            if field == STATUS_FIELD:
                if _norm(value) in _APPROVAL_STATUSES:
                    row_error = (f"{where}: Status {value!r} is an approval state — Cowork "
                                 "may only create rows as a working state (Draft / Awaiting Asset)")
                    break
            cerr = _validate_constrained(field, value)
            if cerr:
                row_error = f"{where}: {cerr}"
                break
            row_updates.append((_a1(tab, col, dest), value))
            if field:
                row_fields.add(field)
        if row_error:
            errors.append(row_error)
            continue
        # default new rows to Draft so generate will process them
        if status_col and STATUS_FIELD not in row_fields:
            row_updates.append((_a1(tab, status_col, dest), "Draft"))
        updates.extend(row_updates)
        planned.append({"row_id": row_id, "row": dest, "cells": len(row_updates)})

    result = {"calendar_id": calendar_id, "mode": mode, "planned": len(planned),
              "rows": planned, "errors": errors, "sheet_link": link}
    if errors:
        result["note"] = (f"{len(errors)} invalid row(s) — nothing was written. "
                          "Fix the errors and retry.")
        return result
    if mode == "dry-run":
        result["note"] = f"dry-run: {len(planned)} row(s) would be appended from row {start_row}."
        return result

    from ..core.sheets import SheetsClient
    sheets = SheetsClient(config.credentials_path(), config.token_path())
    res = sheets.batch_update(sid, updates)
    result["updated_cells"] = res.get("totalUpdatedCells", len(updates))
    emit(f"add: appended {len(planned)} row(s) ({result['updated_cells']} cells) to "
         f"{sheet_ops.live_sheet_name(calendar_id)}")
    return result
